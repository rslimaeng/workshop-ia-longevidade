#!/usr/bin/env python3
"""
Gates de auditoria do site do workshop da Longevidade.

Cada gate roda contra todos os .html publicados E é provado contra um defeito
injetado numa cópia em memória. Gate que não acusa o próprio defeito é gate que
não existe, e o script falha por isso, não só por achar problema no site.

Rodar:  python3 _build/gates.py
Saída:  FALHA por gate + exit code. Exit code sozinho nunca é prova: leia a saída.
"""

import os
import re
import sys
import glob

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# A capa pode falar de tempo: a agenda é o que foi contratado. As páginas de
# conteúdo são material do participante e não carregam minutagem.
PAGINAS_COM_AGENDA = {"index.html"}


def paginas():
    """Todo .html publicado, com caminho relativo à raiz do site."""
    achados = []
    for p in glob.glob(os.path.join(RAIZ, "**", "*.html"), recursive=True):
        rel = os.path.relpath(p, RAIZ)
        # _build é fonte, _rascunho é diagrama em teste: nenhum dos dois é página
        # publicada, e varrer os dois faz o gate auditar arquivo que ninguém abre.
        if rel.split(os.sep)[0] in ("_build", "_rascunho"):
            continue
        achados.append((rel, open(p, encoding="utf-8").read()))
    return sorted(achados)


def linhas_com(texto, padrao, flags=re.I):
    """Devolve (nº da linha, trecho) de cada linha inteira que casa. Nunca recorta
    antes de filtrar: recortar antes é como o gate de 'pede nome' deu falso
    positivo em cima de 'Não escreva seu nome'."""
    saida = []
    for n, linha in enumerate(texto.split("\n"), 1):
        if re.search(padrao, linha, flags):
            saida.append((n, linha.strip()[:150]))
    return saida


# Só tags de bloco. strong, em, span e a são inline: quebrar neles partiria a
# frase ao meio e faria o gate de "a fonte está na mesma frase" acusar sozinho.
BLOCOS = ("p", "div", "li", "h1", "h2", "h3", "h4", "td", "th", "tr",
          "section", "summary")


def texto_visivel(html):
    """Só o que o participante lê: sem <style>, <script>, comentários e tags.

    Preserva a quebra entre blocos. Colapsar tudo numa linha só faz o gate
    perder a noção de 'na mesma frase' e devolve a página inteira como
    mensagem de erro, que foi como o G11 nasceu cego.
    """
    s = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.S)
    s = re.sub(r"<script[^>]*>.*?</script>", " ", s, flags=re.S)
    s = re.sub(r"<!--.*?-->", " ", s, flags=re.S)
    # O espaço rígido que o gerador cola entre "na" e "sua" é espaço para quem
    # lê. Se ele chegasse aqui como caractere próprio, todo gate que compara
    # frase passaria a acusar a si mesmo depois da correção de quebra de linha.
    s = s.replace("&nbsp;", " ").replace("\u00a0", " ")
    # Colapsa ANTES de marcar bloco: quebra de linha no arquivo-fonte não é
    # quebra na tela, e tratar as duas como iguais partia a frase da Capgemini.
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"</(" + "|".join(BLOCOS) + r")>", "\n", s)
    s = re.sub(r"<br\s*/?>", "\n", s)
    s = re.sub(r"<[^>]+>", " ", s)
    linhas = [re.sub(r"[ \t]+", " ", l).strip() for l in s.split("\n")]
    return "\n".join(l for l in linhas if l)


# ----------------------------------------------------------------- os gates

def g1_travessao(rel, html):
    """Travessão é proibido em todo material do Rafael."""
    falhas = []
    for n, l in linhas_com(texto_visivel(html), r"—"):
        falhas.append("linha com travessão: " + l)
    return falhas


VOCAB_INTERNO = [
    r"\bonda \d", r"\bgate\b", r"\bhandoff\b", r"\bspec\b", r"\bcanonical\b",
    r"\bsubsistema\b", r"\bfio[s]? em aberto\b", r"\bsubagente\b",
    r"\bclaude code\b", r"\bPCTFL\b", r"\bPOC-1\b", r"\bcommit\b",
    r"\bprompt de sistema\b", r"\bmarkdown\b", r"\bfrontmatter\b",
    # projetos e pessoas de outras frentes do Rafael
    r"\bmaria pitanga\b", r"\bmallory\b", r"\bpouchain\b", r"\bbumpai\b",
    r"\birede\b", r"\btracemat\b", r"\bporanga\b", r"\bcirilo\b",
    r"\bclineu\b", r"\bsamuel\b", r"\baurora\b",
]


def g2_vocabulario_interno(rel, html):
    """Vocabulário interno e nome de outra frente não vão para a tela."""
    vis = texto_visivel(html)
    falhas = []
    for termo in VOCAB_INTERNO:
        for n, l in linhas_com(vis, termo):
            falhas.append("vocabulário interno {}: {}".format(termo, l))
    return falhas


# Premissa que o cliente deu. Devolver isso ocupa espaço sem informar.
PREMISSA_DO_CLIENTE = [
    r"banco de ideias", r"\broadmap\b", r"\bdiretoria\b",
    r"vocês (já )?constru", r"vocês (já )?fizeram", r"que vocês (já )?levantaram",
    r"plataforma de gestão integrada", r"\bgemini\b",
    r"as prioridades que vocês", r"os setores que vocês",
    r"o diagnóstico que vocês", r"as ferramentas que vocês",
    # nomes de dentro do cliente e dos instrutores anteriores
    r"\bizadora\b", r"\bisaac\b", r"\balef\b", r"\bdiego\b",
]


def g3_premissa_devolvida(rel, html):
    """Não devolver ao cliente premissa que ele mesmo deu."""
    vis = texto_visivel(html)
    falhas = []
    for termo in PREMISSA_DO_CLIENTE:
        for n, l in linhas_com(vis, termo):
            falhas.append("premissa devolvida {}: {}".format(termo, l))
    return falhas


def g4_classe_sem_css(rel, html):
    """Toda classe usada no HTML tem regra no CSS da própria página."""
    css = " ".join(re.findall(r"<style[^>]*>(.*?)</style>", html, flags=re.S))
    definidas = set(re.findall(r"\.([A-Za-z][\w-]*)", css))
    usadas = set()
    for attr in re.findall(r'class="([^"]+)"', html):
        for c in attr.split():
            usadas.add(c)
    orfas = sorted(usadas - definidas)
    return ["classe sem CSS: ." + c for c in orfas]


def g5_gabarito_fechado(rel, html):
    """A resposta do caso nunca nasce aberta: quebraria o exercício de 3 passos."""
    falhas = []
    for m in re.finditer(r"<details[^>]*>", html):
        tag = m.group(0)
        if "gabarito" in tag and re.search(r"\bopen\b", tag):
            falhas.append("gabarito nasce aberto: " + tag[:100])
    return falhas


def g6_links_resolvem(rel, html):
    """Todo link relativo aponta para arquivo que existe no disco."""
    base = os.path.dirname(os.path.join(RAIZ, rel))
    falhas = []
    for href in re.findall(r'href="([^"#]+)"', html):
        if href.startswith(("http://", "https://", "mailto:", "data:")):
            continue
        alvo = os.path.normpath(os.path.join(base, href))
        if os.path.isdir(alvo):
            alvo = os.path.join(alvo, "index.html")
        if not os.path.exists(alvo):
            falhas.append("link quebrado: {} (procurei em {})".format(
                href, os.path.relpath(alvo, RAIZ)))
    return falhas


def g7_nav_bate_com_secoes(rel, html):
    """A nav lateral aponta 1:1 para as seções que existem, sem sobra dos dois lados."""
    # Casar 'class="side-nav-item"' inteiro amarra o gate a nao ganhar mais
    # nenhuma classe. Quando o gerador poe p-curto no <li>, o gate acha zero
    # item, devolve [] e passa cego. Procurar pelo efeito.
    nav = re.findall(r'<li[^>]*class="[^"]*\bside-nav-item\b[^"]*"[^>]*>\s*<a href="#([^"]+)"', html)
    secoes = re.findall(r'<section class="block" id="([^"]+)"', html)
    if not nav:
        return []
    falhas = []
    for i in nav:
        if i not in secoes:
            falhas.append("nav aponta para seção inexistente: #" + i)
    for s in secoes:
        if s not in nav:
            falhas.append("seção fora da nav: #" + s)
    return falhas


DIRECAO_DE_CENA = [
    r"pergunte à sala", r"pergunte para a sala", r"a sala responde",
    r"espere o silêncio", r"aguarde o silêncio", r"plano b",
    r"o que apontar", r"roteiro de palco", r"dê um tempo",
    r"circule pela sala", r"anote no quadro", r"projete a tela",
]


def g8_direcao_de_cena(rel, html):
    """Página é do participante, não roteiro do instrutor."""
    vis = texto_visivel(html)
    falhas = []
    for termo in DIRECAO_DE_CENA:
        for n, l in linhas_com(vis, termo):
            falhas.append("direção de cena {}: {}".format(termo, l))
    return falhas



def sem_cartoes_de_video(html):
    """Devolve o HTML sem os blocos <a class="fonte">, para gates que não devem
    ler o cartão de fonte. Recorta só o cartão: o que vem antes e depois dele
    continua sendo auditado."""
    saida, i = [], 0
    while True:
        ini = html.find('<a class="fonte"', i)
        if ini == -1:
            saida.append(html[i:])
            return "".join(saida)
        fim = html.find("</a>", ini)
        if fim == -1:
            saida.append(html[i:])
            return "".join(saida)
        saida.append(html[i:ini])
        i = fim + 4


def g9_minutagem_fora_da_capa(rel, html):
    """Minutagem é controle de condução. Só a capa carrega, porque lá é o contrato."""
    if rel in PAGINAS_COM_AGENDA:
        return []
    # O cartão de fonte em vídeo declara a duração da gravação ("13 minutos"),
    # que é fato do material, igual a "sete páginas", e não diz nada sobre como
    # conduzir o dia. A exceção é ESTRUTURAL, não por vizinhança de palavra:
    # vale só dentro de <a class="fonte"> ... </a>, e para de valer no </a>.
    vis = texto_visivel(sem_cartoes_de_video(html))
    falhas = []
    # A regra protegida é "quanto tempo dura o bloco da aula", que é condução.
    # A duração de um MATERIAL (uma gravação de 48 minutos) é fato do insumo, do
    # mesmo tipo que "sete páginas", e não diz nada sobre como conduzir o dia.
    material = re.compile(
        r"(transcri\u00e7\u00e3o|entrevista|grava\u00e7\u00e3o|\u00e1udio|v\u00eddeo|convers[ao])",
        re.I,
    )
    for padrao in [r"\d+\s*min\b", r"\b\d+h\d{2}\b", r"\b\d+\s*minutos\b"]:
        for n, l in linhas_com(vis, padrao):
            if material.search(l):
                continue
            falhas.append("minutagem fora da capa: " + l)
    return falhas


def g13_botao_copia_tem_alvo(rel, html):
    """Botão de copiar aponta para um id que existe na mesma página.

    Se o id não bate, o botão não faz nada e ninguém percebe: a página abre,
    o botão aparece, e a falha só existe no clique.
    """
    falhas = []
    ids = set(re.findall(r'id="([^"]+)"', html))
    for alvo in re.findall(r'data-copia="([^"]+)"', html):
        if alvo not in ids:
            falhas.append("botão de copiar aponta para id inexistente: " + alvo)
    return falhas


def g14_cem_pontos_batem_com_o_texto(rel, html):
    """A figura dos cem quadrados tem que contar a mesma coisa que a frase.

    A figura é a prova visual do número. Se a frase diz 13 e a figura acende
    outro tanto, a página mente em dois lugares ao mesmo tempo.
    """
    if 'class="cem-grade"' not in html:
        return []
    falhas = []
    grade = re.search(r'<div class="cem-grade"[^>]*>(.*?)</div>', html, flags=re.S)
    if not grade:
        return ["figura dos cem pontos sem grade legível"]
    corpo = grade.group(1)
    acesos = len(re.findall(r'class="cem-p aceso"', corpo))
    total = len(re.findall(r'class="cem-p', corpo))
    if total != 100:
        falhas.append("a figura tem {} quadrados, e deveria ter 100".format(total))
    vis = texto_visivel(html)
    declarado = re.search(r"(\d+)%\s*dos projetos", vis)
    if declarado and int(declarado.group(1)) != acesos:
        falhas.append(
            "a frase diz {}% e a figura acende {} quadrados".format(
                declarado.group(1), acesos))
    if not declarado:
        falhas.append("figura dos cem pontos sem o número declarado no texto")
    return falhas


def g15_prompt_baixado_bate_com_a_tela(rel, html):
    """O .md que a pessoa baixa é igual ao prompt que ela lê na tela.

    Baixar e copiar têm que entregar a mesma coisa. Se divergirem, quem baixou
    trabalha com uma versão que ninguém revisou.
    """
    import html as _html
    falhas = []
    base = os.path.dirname(os.path.join(RAIZ, rel))
    for m in re.finditer(
        r'<div class="prompt-conteudo" id="([^"]+)">(.*?)</div>', html, flags=re.S
    ):
        nome = m.group(1).replace("prompt-", "") + ".md"
        caminho = os.path.join(base, nome)
        if not os.path.exists(caminho):
            falhas.append("prompt sem arquivo para baixar: " + nome)
            continue
        na_tela = _html.unescape(m.group(2)).strip()
        no_arquivo = open(caminho, encoding="utf-8").read().strip()
        if na_tela != no_arquivo:
            falhas.append(
                "o {} baixado difere do que está na tela ({} contra {} caracteres)".format(
                    nome, len(no_arquivo), len(na_tela)))
    return falhas


def g16_campos_do_canvas_com_piso(rel, html):
    """As caixas lado a lado do canvas têm piso de altura.

    Sem o piso o seletor sai 3px mais baixo que o campo de texto ao lado, porque
    o navegador dá ao seletor uma altura interna própria. Medido: 47 e 44 antes,
    47 e 47 depois. Rafael apontou isso olhando a tela, antes de qualquer gate.
    """
    if "canvas" not in rel:
        return []
    css = " ".join(re.findall(r"<style[^>]*>(.*?)</style>", html, flags=re.S))
    if not re.search(r"input\.txt\s*,\s*select\.txt\s*\{[^}]*min-height", css):
        return ["campos do canvas sem piso de altura: as caixas lado a lado saem diferentes"]
    return []


# Os cinco pilares são vocabulário fechado: a régua do celular, a folha impressa
# e a proposta aprovada usam exatamente estes rótulos.
PILARES = [
    "O que não vira registro não existe",
    "Resultado ruim é falta de contexto, não de capacidade",
    "Quebrar em fases e conferir entre elas",
    "Toda correção vira regra",
    "Nada começa sem quatro campos preenchidos",
]


# Onde o rótulo de um pilar é AFIRMADO. Perguntar "existe na página?" deixava
# passar a paráfrase enquanto sobrasse uma cópia certa em qualquer outro canto:
# quando a capa ganhou o ciclo, o rótulo passou a aparecer duas vezes e o gate
# parou de acusar o defeito injetado. Presença não é conferência.
ROTULOS_DE_PILAR = ("pilar-t", "loop-no-t", "fp-t")


def g10_pilares_literais(rel, html):
    """Cada lugar que nomeia um pilar usa o rótulo canônico, sem paráfrase."""
    if rel not in ("index.html", "pilares/index.html"):
        return []
    falhas = []
    achados = []
    for classe in ROTULOS_DE_PILAR:
        # class= tolerante: o gerador acrescenta classe, e casar o atributo
        # inteiro deixaria este gate sem alvo nenhum, ou seja, cego
        for m in re.finditer(
            r'<div class="[^"]*\b' + classe + r'\b[^"]*">(.*?)</div>', html, flags=re.S
        ):
            t = re.sub(r"<[^>]+>", "", m.group(1))
            # compara o que se LÊ: o espaço rígido da cola é um espaço na tela,
            # e o rótulo continua sendo o canônico
            t = re.sub(r"\s+", " ", t.replace("&nbsp;", " ").replace("\u00a0", " ")).strip()
            achados.append((classe, t))

    for classe, t in achados:
        if t not in PILARES:
            falhas.append(
                "rótulo de pilar reescrito em .{}: {!r}".format(classe, t[:80]))

    # e os cinco continuam tendo que existir na página
    vis = texto_visivel(html)
    for pilar in PILARES:
        if pilar not in vis:
            falhas.append("pilar ausente: " + pilar)
    return falhas


def g11_capgemini_com_fonte(rel, html):
    """Número de terceiro entra com a fonte colada, nunca solto."""
    vis = texto_visivel(html)
    falhas = []
    for n, l in linhas_com(vis, r"\b13%|\b87 de cada"):
        if not re.search(r"capgemini", l, re.I):
            falhas.append("número de terceiro sem a fonte na mesma frase: " + l)
    return falhas


def g12_contagem_dos_pilares(rel, html):
    """Onde o texto diz cinco pilares, existem cinco blocos de pilar."""
    if rel != "index.html":
        return []
    falhas = []
    # A capa mostra os pilares como ciclo. Os blocos soltos saíram porque
    # repetiam o que o ciclo já diz, então quem conta agora são as etapas.
    ciclo = len(re.findall(r'<div class="loop-no">', html))
    if ciclo != 5:
        falhas.append("a capa diz cinco pilares e o ciclo desenha {} etapas".format(ciclo))
    # cada etapa aponta onde o pilar é provado: sem isso a capa deixa de ser índice
    selos = len(re.findall(r'<span class="loop-no-onde">', html))
    if selos != ciclo:
        falhas.append("{} etapas no ciclo e {} selos de onde é provado".format(ciclo, selos))
    return falhas


# id, o que confere, função, defeito que prova o gate, onde injetar o defeito

# ---------------------------------------------------------------- G17
# Caractere de outro alfabeto que renderiza parecido com o latino. Eu escrevi
# "потencial" com dois caracteres cirílicos e o texto passou por cima de todos
# os outros gates: na tela ele é indistinguível. Só a checagem do alfabeto pega.
NAO_LATINO = re.compile(r"[\u0400-\u04FF\u0370-\u03FF\u0500-\u052F]")


def g17_alfabeto(rel, html):
    falhas = []
    for n, l in linhas_com(texto_visivel(html), NAO_LATINO.pattern, flags=0):
        achados = NAO_LATINO.findall(l)
        falhas.append("caractere de outro alfabeto ({}) na linha {}: {}".format(
            "".join(sorted(set(achados))), n, l[:90]))
    return falhas


# ---------------------------------------------------------------- G18
# Os sete números "X de 22" vêm do diagnóstico prévio que o cliente respondeu.
# Se a planilha ganhar respostas, a página mente e ninguém percebe: o insumo é
# a fonte da verdade, e quem descreve o insumo tem que ser conferido contra ele.
PLANILHA = os.path.join(
    os.path.dirname(RAIZ), "2-insumos-do-cliente", "briefing",
    "Diagnóstico prévio — IA no Back Office (respostas).xlsx")

OPCOES_DIAGNOSTICO = [
    "Organizar e resumir grandes volumes de informação",
    "Interpretar documentos, contratos e legislações",
    "Apoiar decisões com mais agilidade e embasamento",
    "Montar relatórios e apresentações para gestores",
    "Analisar dados, planilhas e indicadores de resultado",
    "Fazer cálculos e simular cenários (custos, impostos, margens)",
    "Automatizar tarefas repetitivas e rotinas administrativas",
]


def _contagem_real():
    """Lê a planilha do cliente e devolve {opção: quantas marcaram}, e o total.
    Devolve None se a planilha não estiver acessível: gate que depende de insumo
    externo tem que dizer que não conferiu, e não passar calado."""
    try:
        import openpyxl
    except ImportError:
        return None
    if not os.path.exists(PLANILHA):
        return None
    w = openpyxl.load_workbook(PLANILHA, data_only=True)
    s = w[w.sheetnames[0]]
    linhas = [r for r in list(s.iter_rows(values_only=True))[1:] if r[0]]
    cont = {}
    for o in OPCOES_DIAGNOSTICO:
        cont[o] = sum(1 for r in linhas if o in str(r[3] or ""))
    return cont, len(linhas)


def g18_numeros_do_diagnostico(rel, html):
    """As sete atividades da página continuam sendo as do insumo, e cada uma
    aponta o pilar que o próprio card diz que falta.

    A página deixou de exibir os números do levantamento a pedido do Rafael: ele
    é de outra turma e de outro momento, e citá-lo aqui prometia à sala uma
    pesquisa que não foi feita para este dia. O insumo continua sendo a fonte das
    sete atividades, então o gate continua conferindo contra ele — o que mudou é
    que a conferência virou de proveniência, e não de número na tela.
    """
    if rel != "ai-first/index.html":
        return []
    falhas = []
    vis = texto_visivel(html)

    # 1. o levantamento não é mais citado na tela
    for termo in ("diagnóstico", "responderam", "de 22", "Vinte e duas"):
        if termo.lower() in vis.lower():
            for n, l in linhas_com(vis, re.escape(termo)):
                falhas.append("o levantamento de outra turma voltou para a tela: " + l[:90])

    # 2. as sete atividades da página são as sete do insumo
    cards = re.findall(CL("ex-t") + r"[^>]*>([^<]+)<", html)
    real = _contagem_real()
    if real is None:
        falhas.append("não foi possível conferir as atividades contra a planilha do cliente")
    else:
        for t in cards:
            t = t.replace("&nbsp;", " ").strip()
            if not [o for o in OPCOES_DIAGNOSTICO if o.startswith(t[:28])]:
                falhas.append("atividade que não existe no insumo: " + t)
        if len(cards) != len(OPCOES_DIAGNOSTICO):
            falhas.append("a página lista {} atividades e o insumo tem {}".format(
                len(cards), len(OPCOES_DIAGNOSTICO)))

    # 3. o rótulo de cada card diz o mesmo pilar que o corpo dele diz que falta
    blocos = re.split(CL("ex") + ">", html)[1:]
    for b in blocos:
        rot = re.search(CL("ex-area") + r"[^>]*>([^<]+)<", b)
        tit = re.search(CL("ex-t") + r"[^>]*>([^<]+)<", b)
        if not rot or not tit:
            continue
        no_rotulo = set(re.findall(r"\d+", rot.group(1)))
        corpo = texto_visivel(b)
        falta = re.search(r"O que falta\s*(.{0,90})", corpo, re.S)
        if not falta:
            falhas.append("card sem 'O que falta': " + tit.group(1)[:50])
            continue
        # "Pilar 1 e pilar 3" repete a palavra no segundo, e o par não vem junto
        no_corpo = set()
        for m in re.finditer(r"[Pp]ilar(?:es)?\s+(\d)(?:\s+e\s+(?:[Pp]ilar\s+)?(\d))?",
                             falta.group(1)):
            no_corpo.update(g for g in m.groups() if g)
        if not no_corpo:
            falhas.append("o card {!r} não nomeia pilar nenhum em 'O que falta'".format(
                tit.group(1)[:40]))
        elif no_rotulo != no_corpo:
            falhas.append("o card {!r} tem rótulo {} e o corpo fala do pilar {}".format(
                tit.group(1)[:40], sorted(no_rotulo), sorted(no_corpo)))
    return falhas


# ---------------------------------------------------------------- G19
# A régua mede a mesma coisa nas duas pontas. Se a abertura e o fechamento
# divergirem em número de alternativas ou no texto delas, a comparação de
# antes e depois compara coisas diferentes e a demonstração ao vivo mente.
def g19_regua_simetrica(rel, html):
    if rel != "canvas/index.html":
        return []
    falhas = []

    def alts(nome):
        return re.findall(
            r'<input type="radio" name="' + nome + r'" value="(\d+)"><span>([^<]*)</span>',
            html)

    for i in range(1, 6):
        a, f = alts("a%d" % i), alts("f%d" % i)
        if len(a) != 3:
            falhas.append("pergunta a{} tem {} alternativas, e a régua é de 3".format(i, len(a)))
        if len(f) != 3:
            falhas.append("pergunta f{} tem {} alternativas, e a régua é de 3".format(i, len(f)))
        if [t for _, t in a] != [t for _, t in f]:
            falhas.append("a pergunta {} tem texto diferente na abertura e no fechamento".format(i))
        if [v for v, _ in a] != ["1", "2", "3"]:
            falhas.append("os valores de a{} não são 1,2,3: {}".format(i, [v for v, _ in a]))
    return falhas


# ---------------------------------------------------------------- G20
# O passo a passo herdado do outro workshop só acrescenta alguma coisa se cada
# passo disser QUEM faz. Passo sem ator vira lista comum, e o participante lê
# sem saber se aquilo é trabalho dele ou da máquina.
def g20_passo_tem_ator(rel, html):
    falhas = []
    for m in re.finditer("<div " + CL("fl") + r">(.*?)</div>\s*</div>", html, flags=re.S):
        bloco = m.group(1)
        if not re.search(r'class="[^"]*\bfl-ator\b', bloco):
            t = re.search(r'<div class="fl-t">([^<]*)', bloco)
            falhas.append("passo sem ator nomeado: " + (t.group(1)[:70] if t else "?"))
    return falhas


# ---------------------------------------------------------------- G21
# A legenda da figura afirma que as caixas iniciais dos dois loops são iguais,
# e é essa igualdade que faz o argumento: o que muda é só a volta. Se alguém
# editar um lado, a legenda passa a mentir sobre a própria figura.
def g21_loops_comecam_igual(rel, html):
    if rel != "index.html":
        return []
    # Fatiar por posição, e não por regex com .*?: "lp-rot" também começa com
    # "lp", então o não-guloso fechava o bloco na primeira linha e devolvia
    # zero etapas num HTML correto.
    ini_aberto = html.find('<div class="lp">')
    ini_fechado = html.find('<div class="lp fechado">')
    # Procurar por 'class="fig-leg"' inteiro amarra o gate à ordem exata do
    # atributo: acrescentar uma classe ao <p> deixou este gate cego uma vez.
    m_fim = re.search(r'<p [^>]*class="[^"]*\bfig-leg\b', html[ini_fechado:]) if ini_fechado > 0 else None
    fim = ini_fechado + m_fim.start() if m_fim else -1
    if ini_aberto < 0 or ini_fechado < 0 or fim < 0:
        return ["não achei os dois loops desenhados na capa"]
    if not (ini_aberto < ini_fechado < fim):
        return ["os dois loops estão fora de ordem na capa"]

    def etapas(t):
        return re.findall(r'<span class="lp-et">([^<]+)</span>', t)

    aberto = etapas(html[ini_aberto:ini_fechado])
    fechado = etapas(html[ini_fechado:fim])
    if len(aberto) != 3:
        return ["o loop aberto tem {} etapas, e a legenda fala de três".format(len(aberto))]
    if fechado[:3] != aberto:
        return ["as três primeiras etapas dos dois loops divergem: {} contra {}".format(
            aberto, fechado[:3])]
    if len(fechado) <= len(aberto):
        return ["o loop fechado não tem etapa a mais que o aberto"]
    return []


# ---------------------------------------------------------------- G22
# A folha em papel e a régua da tela medem a mesma coisa, e a demonstração ao
# vivo compara as duas na mesma planilha. Se alguém mexer numa e não na outra,
# a comparação passa a somar coisas diferentes sem nenhum erro aparecer.
def g22_papel_bate_com_a_tela(rel, html):
    if rel != "papel/index.html":
        return []
    canvas = os.path.join(RAIZ, "canvas", "index.html")
    if not os.path.exists(canvas):
        return ["não achei a régua da tela para comparar com o papel"]
    c = open(canvas, encoding="utf-8").read()

    da_tela = []
    for i in range(1, 6):
        alts = re.findall(
            r'<input type="radio" name="a%d" value="\d"><span>([^<]+)</span>' % i, c)
        da_tela.append(alts)
    if any(len(a) != 3 for a in da_tela):
        return ["a régua da tela não tem três alternativas em toda pergunta"]

    do_papel = []
    for m in re.finditer(r'<div class="pl-ops">(.*?)</div>', html, flags=re.S):
        do_papel.append(re.findall(r'<span class="pl-tx">([^<]+)</span>', m.group(1)))

    falhas = []
    if len(do_papel) != 10:
        falhas.append("o papel tem {} blocos de pergunta, e a régua tem cinco em cada ponta".format(
            len(do_papel)))
        return falhas
    # Os dois lados carregam o espaço rígido da correção de quebra de linha, mas
    # em codificações diferentes: a tela guarda "&nbsp;" literal e o papel já sai
    # do gerador com o caractere. Comparar sem normalizar acusava dez divergências
    # que ninguém vê na folha. O que se compara é o que o leitor lê.
    import html as _h

    def como_se_le(t):
        return _h.unescape(t).replace("\u00a0", " ").strip()

    for i in range(5):
        tela = [como_se_le(t) for t in da_tela[i]]
        for ponta, bloco in (("abertura", do_papel[i]), ("fechamento", do_papel[i + 5])):
            papel = [como_se_le(t) for t in bloco]
            if papel != tela:
                falhas.append("pergunta {} na {} do papel diverge da tela: {} contra {}".format(
                    i + 1, ponta, papel, tela))
    # o teto impresso tem que bater com três alternativas e cinco perguntas
    vis = texto_visivel(html)
    if "de 5 a 15" not in vis:
        falhas.append("o papel não diz a faixa do total, ou ela não é de 5 a 15")
    return falhas


# ---------------------------------------------------------------- G23
# A numeração das seções virou "04, 05, 06, 04" quando duas seções novas
# entraram no meio com rótulo de texto. O leitor usa esses números para se
# localizar na navegação lateral, e dois "04" na mesma página quebram isso.
# O G7 não pega: para ele a nav e as seções continuavam batendo uma a uma.
def g23_numeracao_das_secoes(rel, html):
    rotulos = re.findall(r'<div class="block-num">([^<]*)</div>', html)
    numerados = [r for r in rotulos if re.match(r"^\d+\s*·", r.strip())]
    if not numerados:
        return []          # página de rótulo textual, como a capa
    if len(numerados) != len(rotulos):
        return ["a página mistura seção numerada com seção sem número: {}".format(
            [r for r in rotulos if r not in numerados])]
    nums = [int(re.match(r"^(\d+)", r.strip()).group(1)) for r in numerados]
    falhas = []
    vistos = set()
    for n in nums:
        if n in vistos:
            falhas.append("número de seção repetido: {:02d} ({})".format(n, nums))
        vistos.add(n)
    if nums != list(range(1, len(nums) + 1)):
        falhas.append("a numeração das seções não é sequencial: {}".format(nums))
    return falhas


# ---------------------------------------------------------------- G24
# A página do papel afirma quantas folhas saem, em dois lugares: o chip do
# topo e a caixa "para imprimir". Quando a régua virou frente e verso, o chip
# continuou dizendo duas. Número que a página afirma sobre ela mesma é
# conteúdo, e conteúdo tem que bater com o que ela tem.
NUMERO_POR_EXTENSO = {"uma": 1, "duas": 2, "três": 3, "quatro": 4, "cinco": 5}


def g24_folhas_declaradas(rel, html):
    if rel not in ("papel/index.html", "pilares/index.html"):
        return []
    reais = len(re.findall(r'<div class="folha">', html))
    if not reais:
        return ["a página não tem nenhuma folha"]
    vis = texto_visivel(html)
    falhas, achou = [], 0
    padrao = r"\b(uma|duas|três|quatro|cinco)\s+(?:folhas?|páginas?)\s*A4"
    for m in re.finditer(padrao, vis, re.I):
        achou += 1
        dito = NUMERO_POR_EXTENSO[m.group(1).lower()]
        if dito != reais:
            falhas.append("a página diz {} folha(s) A4 e tem {}".format(dito, reais))
    if not achou:
        falhas.append("a página imprimível não declara quantas folhas A4 saem")
    return falhas



# ---------------------------------------------------------------- G25
# O cartão do vídeo é o que dá dono e data ao número que a página cita. Se ele
# vier quebrado (sem link, sem alt, sem chamada), a fonte deixa de ser
# conferível e a página volta a afirmar sozinha. Confere CADA cartão, e não a
# presença de um cartão certo em algum canto do arquivo.
PECAS_DO_CARTAO = ("fonte-rot", "fonte-titulo", "fonte-meta", "fonte-cta")
ENVELHECE = re.compile(r"visualiza\u00e7\u00f5es|inscritos|h\u00e1 \d+ (dia|m\u00eas|mes|semana|ano)", re.I)


def cartoes_de_video(html):
    """Devolve o HTML de cada <a class="fonte"> ... </a>, um por cartão."""
    blocos = []
    for m in re.finditer(r'<a class="fonte"', html):
        fim = html.find("</a>", m.start())
        if fim == -1:
            blocos.append(html[m.start():m.start() + 1200])
        else:
            blocos.append(html[m.start():fim + 4])
    return blocos


def g25_cartao_de_video_completo(rel, html):
    falhas = []
    for i, bloco in enumerate(cartoes_de_video(html), 1):
        rotulo = "cartão {} de {}".format(i, rel)
        if not re.search(r'href="https://www\.youtube\.com/watch\?v=[\w-]+"', bloco):
            falhas.append(rotulo + ": sem link para o vídeo no YouTube")
        capa = re.search(r'<img class="fonte-capa"[^>]*>', bloco)
        if not capa:
            falhas.append(rotulo + ": sem a capa do vídeo")
        else:
            alt = re.search(r'alt="([^"]*)"', capa.group(0))
            if not alt or len(alt.group(1).strip()) < 15:
                falhas.append(rotulo + ": a capa não tem alt que descreva o vídeo")
        for peca in PECAS_DO_CARTAO:
            if peca not in bloco:
                falhas.append("{}: falta {}".format(rotulo, peca))
        gasto = ENVELHECE.search(bloco)
        if gasto:
            falhas.append("{}: traz número que envelhece sozinho ({})".format(
                rotulo, gasto.group(0)))
    return falhas


# ---------------------------------------------------------------- G26
# G6 confere href. Ninguém conferia src, e o caminho da imagem muda entre a capa
# (_img/) e as páginas internas (../_img/): é exatamente o tipo de erro que só
# aparece como quadrado vazio na tela de quem abriu.
def g26_imagens_existem(rel, html):
    base = os.path.dirname(os.path.join(RAIZ, rel))
    falhas = []
    for src_ in re.findall(r'<img[^>]+src="([^"]+)"', html):
        if src_.startswith(("http://", "https://", "data:")):
            continue
        alvo = os.path.normpath(os.path.join(base, src_))
        if not os.path.exists(alvo):
            falhas.append("imagem que não existe: {} (procurei em {})".format(
                src_, os.path.relpath(alvo, RAIZ)))
    return falhas



# ---------------------------------------------------------------- G27
# A pagina precisa caber num celular de 375px. Duas coisas garantem isso e as
# duas somem sem aviso: (a) a tabela larga fica dentro de um envelope que rola
# sozinho, (b) os itens do grid .layout podem encolher. Sem (b), o overflow-x
# de (a) nao serve de nada e a pagina inteira passa a rolar de lado, levando o
# h1 e todo paragrafo junto. Medido no navegador: 582px de rolagem sem a
# escapatoria, 375px com ela.
ESCAPATORIA = ".layout > *{min-width:0}"


def g27_cabe_no_celular(rel, html):
    falhas = []
    if 'class="layout"' in html and ESCAPATORIA not in html:
        falhas.append("usa o grid .layout mas nao declara " + ESCAPATORIA
                      + ": a coluna trava na largura do conteudo mais largo")
    for m in re.finditer(r"<table\b", html):
        antes = html[max(0, m.start() - 600):m.start()]
        if antes.rfind('<div class="table-wrap">') <= antes.rfind("</div>"):
            trecho = html[m.start():m.start() + 70].replace("\n", " ")
            falhas.append("tabela fora do envelope que rola: " + trecho)
    return falhas



# ---------------------------------------------------------------- G28
# O Rafael reprovou a quebra de linha duas vezes. A correção é um passo do
# gerador: ele cola a palavra-função na seguinte com espaço rígido nos blocos
# curtos. Passo de geração some em silêncio (basta alguém editar o HTML
# publicado à mão), e o defeito só reaparece na tela. Aqui a prova é a mesma da
# régua em papel: rodar o passo de novo no artefato final não pode mudar nada.
#
# Medido no navegador, 7 páginas x 3 larguras: 0 quebras ruins com a cola,
# 128 sem ela. Maior trecho colado: 24 caracteres, e nenhuma página estoura.
import importlib.util as _il

_spec = _il.spec_from_file_location(
    "gerador_do_site", os.path.join(RAIZ, "_build", "gerar.py"))
_gerador = _il.module_from_spec(_spec)
_spec.loader.exec_module(_gerador)


def g28_quebra_de_linha_tratada(rel, html):
    falhas = []
    if _gerador.cola_quebra_de_linha(html) != html:
        falhas.append("a cola de quebra de linha não está aplicada nesta página: "
                      "rodar o passo do gerador de novo ainda muda o arquivo")
    limite = _gerador.LIMITE_DO_GRUDADO
    # Unidade colada maior que a linha do celular vira rolagem lateral, que é o
    # defeito que esta correção deveria evitar. Mede no HTML, porque o texto
    # visível já trocou o rígido por espaço comum.
    for m in re.finditer(r"(?:[^\s<>]+(?:&nbsp;)){1,}[^\s<>]+", html):
        colado = m.group(0).replace("&nbsp;", " ")
        if "<" in colado or ">" in colado:
            continue
        if len(colado) > limite:
            falhas.append("trecho colado maior que o limite de {}: {!r}".format(
                limite, colado[:60]))
    return falhas


# ---------------------------------------------------------------- G29
def _arvore_de_frases(html):
    """Percorre com pilha e devolve (hosts, spans_sem_host, hosts_em_item_de_flex,
    hosts_vazios). Regex de vizinhança não serve aqui: o host pode ser <div>,
    <span>, <p> ou <li>, e <div>(.*?)</div> fecha na tag errada."""
    fora = _gerador.FORA_DE_ALCANCE | _gerador._classes_de_container(
        "\n".join(_gerador.ESTILO.findall(html)))
    pilha = [{"tag": None, "classes": set(), "host": False, "tem_frase": False}]
    hosts = 0
    sem_host, em_flex, vazios = [], [], []
    pos = 0
    while True:
        m = _gerador.TAG.search(html, pos)
        if not m:
            break
        pos = m.end()
        if m.group(0).startswith("<!--"):
            continue
        tag = m.group("tag").lower()
        if tag in _gerador.LITERAL and not m.group("fecha"):
            f = re.search(r"</\s*%s\s*>" % tag, html[pos:], re.I)
            if f:
                pos += f.end()
            continue
        if tag in _gerador.VAZIA:
            continue
        if m.group("fecha"):
            fundo = next((k for k in range(len(pilha) - 1, 0, -1)
                          if pilha[k]["tag"] == tag), None)
            if fundo is None:
                continue
            del pilha[fundo + 1:]
            q = pilha.pop()
            if q["host"]:
                hosts += 1
                if not q["tem_frase"]:
                    vazios.append(q["texto"][:60])
            if q["tem_frase"]:
                pilha[-1]["tem_frase"] = True
            continue
        classes = set(_gerador._classes_de(m.group("attrs")))
        ehFrase = m.group(0) == _gerador.SPAN_FRASE
        if ehFrase and not any(q["host"] for q in pilha):
            sem_host.append(html[m.end():m.end() + 60])
        if "fr-host" in classes and (pilha[-1]["classes"] & fora):
            em_flex.append(" ".join(sorted(classes)))
        pilha.append({"tag": tag, "classes": classes, "host": "fr-host" in classes,
                      "tem_frase": ehFrase,
                      "texto": re.sub(r"<[^>]+>", "", html[m.end():m.end() + 200])})
    return hosts, sem_host, em_flex, vazios


def g29_uma_frase_por_linha(rel, html):
    """Cada frase de bloco largo tem a sua linha, e o CSS que faz isso existe."""
    falhas = []
    # 1. O passo recalcula do zero: se rodar de novo e mudar alguma coisa, o
    #    arquivo publicado não é o que o gerador produz. Vale para bloco que
    #    ficou sem cortar, span a mais, span a menos e marca fora do lugar.
    if _gerador.uma_frase_por_linha(html) != html:
        falhas.append("o corte em frases não está aplicado nesta página: rodar o "
                      "passo do gerador de novo ainda muda o arquivo")
    if '<span class="fr">' not in html:
        return falhas
    hosts, sem_host, em_flex, vazios = _arvore_de_frases(html)
    # 2. Frase sem contêiner acima nunca vira bloco: o markup fica inerte.
    for t in sem_host:
        falhas.append("frase cortada sem nenhum fr-host acima: " + t)
    # 3. container-type:inline-size zera a largura de quem se dimensiona pelo
    #    conteúdo. Num item de flex o bloco foi para 0px e o texto saiu uma
    #    palavra por linha. A marca tem de morar em quem tem largura definida.
    for c in em_flex:
        falhas.append("fr-host em item de flex ou grid, a largura colapsa: " + c)
    for t in vazios:
        falhas.append("fr-host sem nenhuma frase dentro: " + t)
    # 4. A outra metade do mecanismo: sem o container query o span não faz nada.
    css = re.sub(r"\s+", "", html)
    if not re.search(r"@container\([^)]*min-width:600px\)\{\.fr\{display:block", css):
        falhas.append("o @container que transforma .fr em bloco não está no CSS "
                      "desta página: os spans não fazem nada")
    return falhas


# O gerador cola espaço rígido no meio das frases, e onde ele cola muda quando o
# conteúdo muda. Injetor que casa literal com espaço comum fica sem alvo e o
# gate passa cego sem que nada no site tenha quebrado: aconteceu com o G11, o
# G21, o G22, o G23, o G25 e o G28. Todo defeito injetado casa por aqui.
ESPACO = r"(?:&nbsp;|\s)+"


def como_escrito(txt):
    """Padrão que casa o texto com ou sem espaço rígido em qualquer espaço."""
    return ESPACO.join(re.escape(p) for p in txt.split())


# ---------------------------------------------------------------- G30
NUM_SEM_ACENTO = {"um": 1, "dois": 2, "tres": 3, "quatro": 4, "cinco": 5,
                      "seis": 6, "sete": 7, "oito": 8, "nove": 9, "dez": 10,
                      "onze": 11, "doze": 12}


# Acima de doze o texto do site também escreve por extenso. Sem estes, um
# número maior que o teto da figura passava despercebido: o gate só reconhecia
# até dezessete e "dezenove itens" não acusava nada.
NUM_ALTO = {"treze": 13, "quatorze": 14, "catorze": 14, "quinze": 15,
            "dezesseis": 16, "dezessete": 17, "dezoito": 18, "dezenove": 19,
            "vinte": 20, "vinte e um": 21, "trinta": 30}


def _sem_acento(t):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", t)
                   if unicodedata.category(c) != "Mn")


def g30_auxiliares_batem_com_a_lista(rel, html):
    """O número de auxiliares escrito no texto bate com o que está listado."""
    if "auxiliares" not in texto_visivel(html):
        return []
    falhas = []
    cartoes = []
    # Fatiar entre um rótulo e o seguinte. Casar ".*?</div></div>" parecia certo
    # e engolia o cartão de baixo: o "Cinco" somou 11, que é o total dos dois.
    marcas = [(m.start(), m.group(1)) for m in re.finditer(
        r'<div class="[^"]*card-eyebrow[^"]*">([^<]*)</div>', html)]
    for i, (pos, rot) in enumerate(marcas):
        fim = marcas[i + 1][0] if i + 1 < len(marcas) else len(html)
        corpo = html[pos:fim]
        # e para na PRIMEIRA lista: sendo o ultimo rotulo da pagina, o corpo ia
        # ate o fim do arquivo e somava <li> de outras secoes
        u = corpo.find("</ul>")
        if u == -1:
            continue
        corpo = corpo[:u]
        n = re.match(r"\s*([A-Za-z\u00c0-\u00ff]+)", _sem_acento(rot).lower())
        if not n or n.group(1) not in NUM_SEM_ACENTO:
            continue
        itens = len(re.findall(r"<li\b", corpo))
        if not itens:
            continue
        dito = NUM_SEM_ACENTO[n.group(1)]
        cartoes.append((rot.strip(), dito, itens))
        if dito != itens:
            falhas.append("o cartão '{}' promete {} e lista {}".format(
                rot.strip(), dito, itens))
    if not cartoes:
        return falhas
    # e o total escrito na frase bate com a soma dos cartões
    total = sum(i for _, _, i in cartoes)
    vis = _sem_acento(texto_visivel(html)).lower()
    # CADA ocorrencia, nao a primeira: o total e afirmado no titulo e no
    # paragrafo, e conferir so uma deixa a outra mentir sozinha
    for m in re.finditer(r"\b([a-z]+) auxiliares\b", vis):
        n = NUM_SEM_ACENTO.get(m.group(1))
        if n is not None and n != total:
            falhas.append("o texto diz {} auxiliares e os cartões somam {}: {!r}".format(
                n, total, vis[max(0, m.start() - 30):m.end() + 10]))
    return falhas


# ---------------------------------------------------------------- G31
# A régua por fase é uma PROVA: ela afirma que a exigência sobe enquanto o
# trabalho cresce e cai quando ele acaba. Prova não se confere pelo texto ao
# lado, se confere pela geometria — se a barra do 14 ficar menor que a do 7, a
# figura passa a dizer o contrário do parágrafo e nenhum gate de texto vê.
RE_RECT = re.compile(r'<rect\s[^>]*x="([\d.]+)"[^>]*y="([\d.]+)"[^>]*width="([\d.]+)"'
                     r'[^>]*height="([\d.]+)"[^>]*>')
RE_TEXT = re.compile(r'<text\s[^>]*x="([\d.]+)"[^>]*>([^<]*)</text>')


def _svg_da_regua(html):
    for m in re.finditer(r"<svg\b.*?</svg>", html, re.S):
        if "régua por fase" in m.group(0):
            return m.group(0)
    return None


def g31_regua_por_fase(rel, html):
    """Na figura da régua, a altura de cada barra acompanha o número que ela diz."""
    svg = _svg_da_regua(html)
    if svg is None:
        return []
    falhas = []
    rot = re.search(r'aria-label="([^"]*)"', svg)
    rotulo = rot.group(1) if rot else ""

    barras = []
    for x, y, w, alt in RE_RECT.findall(svg):
        x, w, alt = float(x), float(w), float(alt)
        if w > 400:          # o fundo da figura, não é barra
            continue
        barras.append((x + w / 2.0, alt))

    numeros = []
    for x, txt in RE_TEXT.findall(svg):
        t = txt.strip()
        if re.fullmatch(r"\d+(-\d+)?", t):
            numeros.append((float(x), t))

    pares = []
    for cx, alt in barras:
        casa = [t for x, t in numeros if abs(x - cx) < 3]
        if not casa:
            falhas.append("barra em x={} sem número escrito em cima".format(cx))
            continue
        pares.append((casa[0], alt))

    if len(pares) < 7:
        falhas.append("a régua tem {} barras com número e deveria ter 7".format(len(pares)))

    for rotulo_num, _ in pares:
        # "15-16" na figura e "15 a 16" no aria-label são o mesmo par de números:
        # o que o gate exige é que os NÚMEROS estejam declarados, não o hífen
        padrao = r"\b" + r"\s*(?:-|a)\s*".join(rotulo_num.split("-")) + r"\b"
        if not re.search(padrao, rotulo):
            falhas.append("o número {} está na figura e não está declarado no aria-label"
                          .format(rotulo_num))

    def teto(t):
        return int(t.split("-")[-1])

    # a geometria não pode contradizer os números: barra maior = número maior
    ordenadas = sorted(pares, key=lambda p: p[1])
    valores = [teto(t) for t, _ in ordenadas]
    if valores != sorted(valores):
        falhas.append("a altura das barras não acompanha os números: "
                      "por altura crescente sai {}".format(valores))

    # e o número que o texto usa como teto é o teto da figura
    maior = max(teto(t) for t, _ in pares) if pares else 0
    vis = _sem_acento(texto_visivel(html)).lower()
    for m in re.finditer(r"\b([a-z]+) itens\b", vis):
        n = NUM_SEM_ACENTO.get(m.group(1))
        if n is None:
            n = NUM_ALTO.get(m.group(1))
        if n is not None and n > maior:
            falhas.append("o texto fala em {} itens e o teto da figura é {}: {!r}"
                          .format(n, maior, vis[max(0, m.start() - 40):m.end() + 10]))
    return falhas


# ---------------------------------------------------------------- G32
# A lista de papéis promete um número na frase acima dela ("Cinco das doze") e
# numera cada linha. Se alguém tirar ou acrescentar uma linha, a frase e os
# rótulos passam a mentir em silêncio — é o mesmo defeito do G30, noutra peça.
def CL(nome):
    """Casa a classe como classe inteira, e não como pedaço de outra.

    \b casa antes do hífen, então r"\bpapel\b" acertava papel-nome e papel-d
    junto: o gate contou 16 linhas onde havia 5. A classe tem que estar
    delimitada por espaço ou pela aspa.
    """
    return r'class="(?:[^"]*\s)?' + re.escape(nome) + r'(?:\s[^"]*)?"'


def g32_papeis_batem_com_a_frase(rel, html):
    """A lista de papéis tem o tamanho que a frase promete, numerada em ordem."""
    # classe casada por regex e não por literal: o gerador acrescenta fr-host
    # na mesma tag, e "class=\"papeis\"" literal deixou este gate cego na
    # primeira rodada. Terceira vez que este defeito aparece.
    if not re.search(CL("papeis"), html):
        return []
    falhas = []
    ini = re.search("<div " + CL("papeis") + ">", html)
    corpo = html[ini.end():]
    fim = corpo.find("<p")               # a lista termina no parágrafo seguinte
    if fim != -1:
        corpo = corpo[:fim]
    linhas = re.split("<div " + CL("papel") + ">", corpo)[1:]
    itens = len(linhas)

    for i, linha in enumerate(linhas, start=1):
        nome = re.search(CL("papel-nome") + r"[^>]*>(.*?)</div>", linha, re.S)
        desc = re.search(CL("papel-d") + r"[^>]*>(.*?)</div>", linha, re.S)
        if not nome or not desc:
            falhas.append("o {}º papel não tem nome e descrição".format(i))
            continue
        cru = re.sub(r"<[^>]+>", " ", nome.group(1))
        m = re.search(r"frente\s+(\d+)", cru)
        if not m:
            falhas.append("o {}º papel não diz de que frente é: {!r}".format(i, cru.strip()))
        elif int(m.group(1)) != i:
            falhas.append("o {}º papel está rotulado como frente {}".format(i, m.group(1)))
        titulo = re.sub(r"frente\s+\d+", "", cru).strip()
        if len(titulo) < 4:
            falhas.append("o {}º papel não tem nome legível".format(i))
        if len(re.sub(r"<[^>]+>", " ", desc.group(1)).split()) < 8:
            falhas.append("o {}º papel não diz para que serve o registro ali".format(i))

    # a frase que anuncia a lista diz quantos são
    vis = _sem_acento(texto_visivel(html)).lower()
    m = re.search(r"\b([a-z]+) das doze\b", vis)
    if not m:
        falhas.append("a lista de papéis não é anunciada por uma frase que diga quantos são")
    else:
        dito = NUM_SEM_ACENTO.get(m.group(1))
        if dito is None:
            falhas.append("a frase anuncia {!r}, que não é um número".format(m.group(1)))
        elif dito != itens:
            falhas.append("a frase promete {} papéis e a lista tem {}".format(dito, itens))
    return falhas


# ---------------------------------------------------------------- G33
# Classe de defeito encontrada NO AR: inserir uma seção renumera as seguintes, e
# as referências cruzadas continuam apontando para o número velho. Duas frases da
# página do conceito mandavam "releia a seção 04" e "o exercício da seção 05"
# quando os blocos eram outros, e nenhum gate via.
RE_REF = re.compile(r"\b(?:d[ao]s?\s+)?(?:seção|secao|bloco)\s+(\d{2})\b", re.I)
RE_BLOCO_NUM = re.compile(r'<div class="[^"]*\bblock-num\b[^"]*">\s*(\d{2})\s*·\s*([^<]*)</div>')


def g33_referencia_a_bloco_existe(rel, html):
    """Referência a 'seção NN' aponta para um bloco que existe, e é o certo."""
    blocos = {n: t.strip() for n, t in RE_BLOCO_NUM.findall(html)}
    if not blocos:
        return []
    falhas = []
    rotulos = {_sem_acento(t.replace("&nbsp;", " ")).lower().strip(): n
               for n, t in blocos.items()}
    vis = texto_visivel(html)
    for m in RE_REF.finditer(vis):
        n = m.group(1)
        if n not in blocos:
            falhas.append("aponta para a seção {} e esta página vai até a {}: {!r}".format(
                n, max(blocos), vis[max(0, m.start() - 45):m.end() + 10].strip()))
            continue
        # e o assunto citado na frase tem que ser o assunto daquele bloco
        # só a frase, e não o que vem antes dela: o texto visível traz o rótulo
        # do próprio bloco logo acima, e a janela larga acusava a si mesma
        antes = _sem_acento(vis[max(0, m.start() - 60):m.start()]).lower()
        antes = antes[antes.rfind("\n") + 1:]
        for rot, dono in rotulos.items():
            if len(rot) < 5:
                continue
            if re.search(r"\b" + re.escape(rot) + r"\b", antes) and dono != n:
                falhas.append("a frase fala de {!r}, que é o bloco {}, e manda ir para o {}: {!r}"
                              .format(rot, dono, n,
                                      vis[max(0, m.start() - 45):m.end() + 10].strip()))
    return falhas


# ---------------------------------------------------------------- G34
# A escada é PROVA de que o custo cresce degrau a degrau: a caixa fica mais alta
# à direita e todas apoiam na mesma base. Se a geometria deixar de crescer, a
# figura passa a dizer que os quatro custam igual, que é o contrário do texto.
def g34_escada_cresce(rel, html):
    """Na figura da escada, a caixa cresce a cada degrau e todas têm a mesma base."""
    svg = None
    for m in re.finditer(r"<svg\b.*?</svg>", html, re.S):
        if "Escada de quatro degraus" in m.group(0):
            svg = m.group(0)
    if svg is None:
        return []
    falhas = []
    caixas = []
    for x, y, w, alt in RE_RECT.findall(svg):
        x, y, w, alt = float(x), float(y), float(w), float(alt)
        if w > 400:
            continue
        caixas.append((x, y, w, alt))
    caixas.sort()

    rotulos = re.findall(r">\s*DEGRAU (\d)\s*<", svg)
    if rotulos != ["1", "2", "3", "4"]:
        falhas.append("os degraus estão rotulados {} e deveriam ser 1, 2, 3 e 4 da esquerda "
                      "para a direita".format(rotulos))
    if len(caixas) != 4:
        falhas.append("a escada tem {} caixas e deveria ter 4".format(len(caixas)))
        return falhas

    alturas = [c[3] for c in caixas]
    if alturas != sorted(alturas) or len(set(alturas)) != 4:
        falhas.append("a altura não cresce a cada degrau: {}".format(alturas))
    bases = {round(c[1] + c[3]) for c in caixas}
    if len(bases) != 1:
        falhas.append("os degraus não apoiam na mesma base: {}".format(sorted(bases)))

    # e o texto que promete o número de alturas bate com o desenho
    vis = _sem_acento(texto_visivel(html)).lower()
    for m in re.finditer(r"\b([a-z]+) (?:alturas|degraus)\b", vis):
        n = NUM_SEM_ACENTO.get(m.group(1)) or NUM_ALTO.get(m.group(1))
        if n is not None and n != len(caixas):
            falhas.append("o texto fala em {} degraus e a figura tem {}: {!r}".format(
                n, len(caixas), vis[max(0, m.start() - 35):m.end() + 8]))
    return falhas


# ---------------------------------------------------------------- G35
# .callout sozinho não pinta nada: a cor e a borda vêm da variante. Um callout
# sem variante nasce invisível como caixa e o leitor vê o texto solto no meio da
# página. Aconteceu na página do conceito e só apareceu quando o Rafael olhou.
def g35_callout_tem_variante(rel, html):
    """Todo callout declara a variante que o pinta."""
    falhas = []
    for m in re.finditer(r'<div class="([^"]*\bcallout\b[^"]*)"', html):
        classes = m.group(1).split()
        if "callout" not in classes:
            continue
        if not {"callout-info", "callout-warn"} & set(classes):
            depois = html[m.end():m.end() + 260]
            t = re.search(r'callout-title[^>]*>([^<]*)', depois)
            falhas.append("callout sem variante de cor, então sem caixa: {!r}".format(
                (t.group(1) if t else depois[:60]).strip()[:70]))
    return falhas


# ---------------------------------------------------------------- G36
# Girar a seta inteira com transform estica a linha do ::before pela largura do
# contêiner: empilhada no celular, ela virava uma barra vertical atravessando os
# cartões de cima e de baixo. Estava no ar na capa desde sempre, e só apareceu
# quando o mesmo componente foi repetido numa página mais estreita.
def g36_seta_do_ciclo_nao_estica(rel, html):
    """No modo empilhado, a linha da seta do ciclo é um traço, não uma barra."""
    if ".loop-trilha" not in html:
        return []
    i = html.find(".loop-trilha{flex-direction:column}")
    if i == -1:
        return ["o ciclo não empilha mais no celular: a regra de coluna sumiu"]
    trecho = html[i:i + 900]
    falhas = []
    seta = re.search(r"\.loop-seta\{([^}]*)\}", trecho)
    antes = re.search(r"\.loop-seta::before\{([^}]*)\}", trecho)
    if not seta or not antes:
        return ["o modo empilhado não define mais a seta do ciclo"]
    if "transform:none" not in seta.group(1).replace(" ", ""):
        falhas.append("a seta gira inteira no modo empilhado, e a linha dela vai "
                      "esticar pela largura do cartão: " + seta.group(1)[:70])
    linha = antes.group(1).replace(" ", "")
    if not re.search(r"width:\d+px", linha):
        falhas.append("a linha da seta não tem largura fixa no modo empilhado: " + linha[:70])
    if "left:0" in linha and "right:0" in linha:
        falhas.append("a linha da seta continua indo de ponta a ponta: " + linha[:70])
    return falhas


GATES = [
    ("G1", "travessão", g1_travessao,
     lambda h: h.replace("<h2>", "<h2>defeito — injetado ", 1), None),
    ("G2", "vocabulário interno na tela", g2_vocabulario_interno,
     lambda h: h.replace("<h2>", "<h2>a onda 3 do handoff ", 1), None),
    ("G3", "premissa do cliente devolvida", g3_premissa_devolvida,
     lambda h: h.replace("<h2>", "<h2>como o banco de ideias que vocês construíram ", 1), None),
    ("G4", "classe sem CSS", g4_classe_sem_css,
     lambda h: h.replace('class="card', 'class="classe-inventada card', 1), None),
    ("G5", "gabarito nasce fechado", g5_gabarito_fechado,
     lambda h: re.sub(r'(<details[^>]*\bgabarito\b[^>]*?)>', r'\1 open>', h, count=1),
     "caso-1/index.html"),
    ("G6", "links resolvem", g6_links_resolvem,
     lambda h: h.replace('href="./canvas/"', 'href="./pagina-que-nao-existe/"', 1),
     "index.html"),
    ("G7", "nav bate com as seções", g7_nav_bate_com_secoes,
     lambda h: re.sub(r'(<section[^>]*\bblock\b[^>]*\bid=")s1(")', r'\g<1>s1-renomeada\g<2>',
                      h, count=1),
     "caso-1/index.html"),
    ("G8", "direção de cena", g8_direcao_de_cena,
     lambda h: h.replace("<h2>", "<h2>pergunte à sala e espere o silêncio ", 1), None),
    ("G9", "minutagem fora da capa", g9_minutagem_fora_da_capa,
     lambda h: h.replace("<h2>", "<h2>bloco de 15 min ", 1), "caso-1/index.html"),
    ("G10", "pilares com o rótulo canônico", g10_pilares_literais,
     lambda h: re.sub(como_escrito("Toda correção vira regra"),
                      "Correções viram regras", h, count=1),
     "index.html"),
    # O defeito tira a fonte e deixa o número sozinho. Casa por regex e não por
    # literal, porque o gerador cola espaço rígido no meio de "da Capgemini" e
    # um literal com espaço comum passou a não encontrar nada, deixando o gate
    # cego sem que nada no site tivesse mudado.
    ("G11", "número de terceiro com a fonte", g11_capgemini_com_fonte,
     lambda h: re.sub(r"Segundo pesquisa da(&nbsp;|\s)Capgemini,\s*(<strong>13%)",
                      r"\2", h, count=1),
     "ficha/index.html"),
    ("G16", "as caixas do canvas têm piso", g16_campos_do_canvas_com_piso,
     lambda h: h.replace("input.txt, select.txt { min-height: 47px; }", "", 1),
     "canvas/index.html"),
    ("G15", "o prompt baixado bate com a tela", g15_prompt_baixado_bate_com_a_tela,
     lambda h: re.sub(como_escrito("# PAPEL"), "# PAPEL ALTERADO", h, count=1),
     "caso-1/index.html"),
    ("G13", "botão de copiar tem alvo", g13_botao_copia_tem_alvo,
     lambda h: h.replace('data-copia="prompt-caso-1"', 'data-copia="prompt-sumiu"', 1),
     "caso-1/index.html"),
    ("G14", "a figura dos cem bate com o texto", g14_cem_pontos_batem_com_o_texto,
     lambda h: re.sub(r'<span class="[^"]*\bcem-p\b[^"]*\baceso\b[^"]*"></span>',
                      '<span class="cem-p"></span>', h, count=1),
     "ficha/index.html"),
    ("G17", "alfabeto latino na tela", g17_alfabeto,
     lambda h: h.replace("registro", "\u0433egistro", 1),
     "ai-first/index.html"),
    ("G18", "as atividades vêm do insumo", g18_numeros_do_diagnostico,
     lambda h: re.sub(como_escrito('<span class="ex-area">pilar 5</span>'),
                      '<span class="ex-area">pilar 4</span>', h, count=1),
     "ai-first/index.html"),
    # O defeito troca o texto de uma alternativa, quebrando a simetria entre as
    # duas pontas. Casa por regex porque o texto carrega espaço rígido depois da
    # correção de quebra de linha, e o literal com espaço comum não achava nada.
    ("G19", "a régua é simétrica e de três", g19_regua_simetrica,
     lambda h: re.sub(r'(<input type="radio" name="f1" value="3"><span>)[^<]+(</span>)',
                      r'\1Fica registrado em algum lugar.\2', h, count=1),
     "canvas/index.html"),
    ("G20", "todo passo diz quem faz", g20_passo_tem_ator,
     lambda h: re.sub(r'<span class="[^"]*\bfl-ator\b[^"]*">Ryan</span>',
                      '<span class="fl-nada">Ryan</span>', h, count=1),
     "ai-first/index.html"),
    ("G21", "os dois loops começam iguais", g21_loops_comecam_igual,
     lambda h: re.sub(como_escrito('<span class="lp-et">Alguém executa</span>'),
                      '<span class="lp-et">Alguem executa</span>', h, count=1),
     "index.html"),
    # O defeito muda o texto de uma alternativa só no papel, e o gate tem que
    # ver que ele deixou de bater com a tela. Regex pelo mesmo motivo do G19:
    # o texto carrega espaço rígido e o literal com espaço comum não acha nada.
    ("G22", "o papel bate com a régua da tela", g22_papel_bate_com_a_tela,
     lambda h: re.sub(r'(<span class="pl-tx">)[^<]*(</span>)',
                      r'\1Texto trocado só no papel.\2', h, count=1),
     "papel/index.html"),
    ("G23", "as seções são numeradas em ordem", g23_numeracao_das_secoes,
     lambda h: re.sub(como_escrito('<div class="block-num">07 ·'),
                      '<div class="block-num">04 ·', h, count=1),
     "ai-first/index.html"),
    ("G24", "as folhas declaradas batem", g24_folhas_declaradas,
     lambda h: re.sub(como_escrito("Três páginas A4"), "Duas páginas A4", h, count=1),
     "papel/index.html"),
    ("G12", "contagem dos pilares", g12_contagem_dos_pilares,
     lambda h: re.sub(r'<div class="([^"]*)\bloop-no\b([^"]*)">',
                      r'<div class="\1loop-no-removido\2">', h, count=1),
     "index.html"),
    ("G25", "o cartão de vídeo está completo", g25_cartao_de_video_completo,
     # regex e nao literal: um espaco rigido no meio da frase deixou este
     # injetor sem alvo, e o gate passou cego sem nada ter mudado no site
     lambda h: re.sub(r'<span class="fonte-cta">.*?</span>', '', h, count=1),
     "ai-first/index.html"),
    ("G26", "as imagens existem no disco", g26_imagens_existem,
     lambda h: h.replace('src="../_img/video-greg-ai-native.jpg"',
                         'src="../_img/capa-que-nao-existe.jpg"', 1),
     "ai-first/index.html"),
    ("G27", "a página cabe no celular", g27_cabe_no_celular,
     lambda h: h.replace(ESCAPATORIA, "", 1),
     "ai-first/index.html"),
    # O defeito tem de cair DENTRO de um <p>: a cola por regex só alcança
    # h1-4, label, <p> e <li>. O primeiro &nbsp; da página mora hoje num
    # <span> do cabeçalho, colado pela outra passagem, e tirá-lo de lá não
    # fazia esta função mudar nada.
    ("G28", "a quebra de linha foi tratada", g28_quebra_de_linha_tratada,
     lambda h: re.sub(r'(<p\b[^>]*>(?:(?!</p>).)*?)&nbsp;', r'\1 ', h, count=1, flags=re.S),
     "index.html"),
    # O defeito desmarca UMA frase, nao todas: um <span> comum no lugar do
    # <span class="fr"> deixa o passo achar que ali falta corte.
    ("G29", "cada frase tem a sua linha", g29_uma_frase_por_linha,
     lambda h: h.replace('<span class="fr">', '<span>', 1),
     "index.html"),
    # O defeito tira UM item de um cartão: o texto continua prometendo seis e a
    # lista passa a ter cinco.
    ("G30", "os auxiliares batem com a lista", g30_auxiliares_batem_com_a_lista,
     lambda h: re.sub(r'(<div class="[^"]*card-eyebrow[^"]*">Seis.*?)<li\b.*?</li>',
                      r'\1', h, count=1, flags=re.S),
     "caso-2/index.html"),
    ("G31", "a régua por fase é verdadeira", g31_regua_por_fase,
     lambda h: re.sub(r'(<rect x="372" y=")162(" width="122" height=")238(")',
                      r'\g<1>330\g<2>70\g<3>', h, count=1),
     "caso-1/index.html"),
    ("G32", "os papéis batem com a frase", g32_papeis_batem_com_a_frase,
     lambda h: h.replace("<span>frente 3</span>", "<span>frente 9</span>", 1),
     "caso-1/index.html"),
    ("G33", "referência a bloco que existe", g33_referencia_a_bloco_existe,
     lambda h: re.sub(como_escrito("o exercício do bloco") + r"(&nbsp;|\s)*09",
                      "o exercício do bloco 05", h, count=1),
     "ai-first/index.html"),
    ("G34", "a escada cresce a cada degrau", g34_escada_cresce,
     lambda h: re.sub(r'(<rect x="580" y=")260(" width="260" height=")310(")',
                      r'\g<1>460\g<2>110\g<3>', h, count=1),
     "ai-first/index.html"),
    ("G35", "callout tem variante de cor", g35_callout_tem_variante,
     lambda h: re.sub(r'<div class="callout callout-(info|warn)"', '<div class="callout"',
                      h, count=1),
     "ai-first/index.html"),
    ("G36", "a seta do ciclo não estica", g36_seta_do_ciclo_nao_estica,
     lambda h: h.replace(".loop-seta{flex:0 0 auto;padding:6px 0;transform:none;height:26px}",
                         ".loop-seta{flex:0 0 auto;padding:6px 0;transform:rotate(90deg)}", 1),
     "index.html"),
]


def main():
    docs = paginas()
    if not docs:
        sys.exit("nenhuma página encontrada em " + RAIZ)

    print("Auditoria do site · {} páginas\n".format(len(docs)))
    print("  " + " · ".join(rel for rel, _ in docs) + "\n")

    total_falhas = 0
    total_checagens = 0
    gates_cegos = []

    for gid, nome, fn, defeito, alvo_defeito in GATES:
        falhas = []
        for rel, html in docs:
            total_checagens += 1
            for f in fn(rel, html):
                falhas.append("{}: {}".format(rel, f))

        # calibração: o gate acusa o próprio defeito?
        alvo = alvo_defeito or docs[0][0]
        html_alvo = dict(docs).get(alvo)
        if html_alvo is None:
            gates_cegos.append("{} · alvo de calibração não existe: {}".format(gid, alvo))
            acusou = False
        else:
            sujo = defeito(html_alvo)
            if sujo == html_alvo:
                gates_cegos.append("{} · o defeito não mudou nada em {}".format(gid, alvo))
                acusou = False
            else:
                acusou = len(fn(alvo, sujo)) > len(fn(alvo, html_alvo))
        if not acusou and html_alvo is not None and dict(docs).get(alvo) is not None:
            if "{} ·".format(gid) not in " ".join(gates_cegos):
                gates_cegos.append("{} · não acusou o defeito injetado em {}".format(gid, alvo))

        marca = "FALHA " if falhas else ("CEGO  " if not acusou else "ok    ")
        print("{} {:<4} {:<38} {} achado(s) · calibrado: {}".format(
            marca, gid, nome, len(falhas), "sim" if acusou else "NÃO"))
        for f in falhas[:12]:
            print("        " + f)
        if len(falhas) > 12:
            print("        ... e mais {}".format(len(falhas) - 12))
        total_falhas += len(falhas)

    print("\n{} checagens · {} achados".format(total_checagens, total_falhas))

    if gates_cegos:
        print("\nGATES QUE NÃO SE PROVARAM (não valem como verificação):")
        for g in gates_cegos:
            print("  " + g)

    if total_falhas or gates_cegos:
        print("\nRESULTADO: FALHA")
        sys.exit(1)
    print("\nRESULTADO: passou")


if __name__ == "__main__":
    main()
